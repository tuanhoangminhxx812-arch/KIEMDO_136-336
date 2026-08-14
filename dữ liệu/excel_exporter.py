import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import re

def export_reconciliation_excel(df_tong_hop, output_filepath, period_text="Tháng 7/2026", cross_matches=None, unmatched_by_pair=None):
    """
    Exports df_tong_hop and discrepancy analysis into a formatted Excel file.
    Sheet 1: TONG_HOP
    Sheet 2: PHAN_TICH_CHENH_LECH (Exact replica of Popup Modal: Side-by-Side Tables)
    """
    wb = openpyxl.Workbook()
    
    # ==========================================
    # SHEET 1: TONG_HOP
    # ==========================================
    ws = wb.active
    ws.title = "TONG_HOP"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles definition
    font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_diff = Font(name="Calibri", size=11, bold=True, color="9C0006")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_diff_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_diff_row = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    thick_bottom = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Title Rows
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Bảng TH Chênh Lệch Từ {period_text}"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    
    ws["E1"] = "CHÊNH LỆCH"
    ws["E1"].font = Font(name="Calibri", size=12, bold=True, color="C00000")
    ws["E1"].alignment = align_center
    
    # Table Header Row 2
    headers = ["HCM", "Điện lực", "Nợ", "Có", "CHÊNH LỆCH"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        if h == "CHÊNH LỆCH":
            cell.fill = fill_diff_header
        else:
            cell.fill = fill_header
            
    # Data Rows (Row 3 onwards)
    start_row = 3
    for idx, row in df_tong_hop.iterrows():
        current_row = start_row + idx
        hcm_val = row['HCM']
        pcvt_val = row['PCVT']
        no_val = row['Nợ']
        co_val = row['Có']
        diff_val = row['CHÊNH LỆCH']
        
        c_hcm = ws.cell(row=current_row, column=1, value=str(hcm_val))
        c_pcvt = ws.cell(row=current_row, column=2, value=str(pcvt_val))
        c_no = ws.cell(row=current_row, column=3, value=no_val)
        c_co = ws.cell(row=current_row, column=4, value=co_val)
        c_diff = ws.cell(row=current_row, column=5, value=f"=C{current_row}-D{current_row}")
        
        c_hcm.alignment = align_center
        c_pcvt.alignment = align_center
        c_no.alignment = align_center
        c_co.alignment = align_center
        c_diff.alignment = align_center
        
        c_hcm.font = font_data
        c_pcvt.font = font_data
        c_no.font = font_data
        c_co.font = font_data
        c_diff.font = font_data
        
        c_no.number_format = '#,##0;(#,##0);"-"'
        c_co.number_format = '#,##0;(#,##0);"-"'
        c_diff.number_format = '#,##0;(#,##0);"-"'
        
        for col_i in range(1, 6):
            ws.cell(row=current_row, column=col_i).border = thin_border
            
        if abs(diff_val) > 0.01:
            c_diff.font = font_diff
            c_diff.fill = fill_diff_row

    # Total Summary Row
    end_data_row = start_row + len(df_tong_hop) - 1
    total_row = end_data_row + 1
    
    ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = font_bold
    ws.cell(row=total_row, column=1).alignment = align_center
    ws.cell(row=total_row, column=2, value="").font = font_bold
    
    c_tot_no = ws.cell(row=total_row, column=3, value=f"=SUM(C3:C{end_data_row})")
    c_tot_co = ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{end_data_row})")
    c_tot_diff = ws.cell(row=total_row, column=5, value=f"=SUM(E3:E{end_data_row})")
    
    for c, val_cell in [(3, c_tot_no), (4, c_tot_co), (5, c_tot_diff)]:
        val_cell.font = font_bold
        val_cell.alignment = align_center
        val_cell.number_format = '#,##0;(#,##0);"-"'
        
    for col_i in range(1, 6):
        cell = ws.cell(row=total_row, column=col_i)
        cell.fill = fill_total
        cell.border = thick_bottom
        
    column_widths = {'A': 15, 'B': 15, 'C': 25, 'D': 25, 'E': 25}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ==========================================
    # SHEET 2: PHAN_TICH_CHENH_LECH (Exact Replica of Web UI Popup Modal: Side-by-Side Tables)
    # ==========================================
    if unmatched_by_pair is not None:
        ws_an = wb.create_sheet(title="PHAN_TICH_CHENH_LECH")
        ws_an.views.sheetView[0].showGridLines = True
        
        # Styles for Sheet 2 matching Web UI Popup Modal
        font_header_main = Font(name="Calibri", size=14, bold=True, color="212529")
        font_pair_header = Font(name="Calibri", size=13, bold=True, color="212529")
        font_diff_label = Font(name="Calibri", size=12, bold=True, color="DC3545")
        font_remark = Font(name="Calibri", size=10.5, color="1E3C72")
        font_sub_tbl_title = Font(name="Calibri", size=11, bold=True, color="212529")
        font_tbl_header = Font(name="Calibri", size=10, bold=True, color="6C757D")
        font_tbl_data = Font(name="Calibri", size=10, color="212529")
        font_success = Font(name="Calibri", size=10, italic=True, color="198754")

        fill_remark_bg = PatternFill(start_color="EBF3FE", end_color="EBF3FE", fill_type="solid")
        fill_tbl_header = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        fill_success_bg = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
        
        border_thin_grid = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        border_remark_box = Border(
            left=Side(style='medium', color='1E3C72'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        ws_an.cell(row=1, column=1, value="📋 BÁO CÁO PHÂN TÍCH CHI TIẾT CHÊNH LỆCH THEO TỪNG CẶP TÀI KHOẢN").font = font_header_main
        
        row_cur = 3
        
        try:
            from data_processor import build_human_remark
        except ImportError:
            def build_human_remark(h, p, d, hm, pm, cm):
                return f"Chênh lệch {d:,.0f} VNĐ do chứng từ chưa hạch toán đồng bộ."

        for pair_key, p_data in unmatched_by_pair.items():
            diff = p_data['diff']
            if abs(diff) <= 0.01:
                continue
                
            h_code = p_data['h_code']
            p_code = p_data['p_code']
            h_missing = p_data['hcm_missing'] # HCM missing
            p_missing = p_data['pcvt_missing'] # Điện lực missing
            
            remark_raw = build_human_remark(h_code, p_code, diff, h_missing, p_missing, cross_matches or [])
            
            # 1. Pair Header Line
            ws_an.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur, end_column=9)
            cell_hdr = ws_an.cell(row=row_cur, column=1, value=f"🎯 Cặp Tài Khoản: {h_code} (HCM) ⇄ {p_code} (Điện lực)")
            cell_hdr.font = font_pair_header
            cell_hdr.alignment = align_left
            row_cur += 1
            
            # 2. Discrepancy Amount Line
            ws_an.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur, end_column=9)
            diff_str = f"Chênh lệch cuối kỳ: {diff:,.0f} VNĐ".replace(",", ".")
            cell_diff = ws_an.cell(row=row_cur, column=1, value=diff_str)
            cell_diff.font = font_diff_label
            cell_diff.alignment = align_left
            row_cur += 1
            
            # 3. Remark Callout Box
            ws_an.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur+1, end_column=9)
            cell_rmk = ws_an.cell(row=row_cur, column=1, value=remark_raw)
            cell_rmk.font = font_remark
            cell_rmk.alignment = align_left
            for r_idx in range(row_cur, row_cur+2):
                for c_idx in range(1, 10):
                    ws_an.cell(row=r_idx, column=c_idx).border = border_remark_box
                    ws_an.cell(row=r_idx, column=c_idx).fill = fill_remark_bg
            ws_an.row_dimensions[row_cur].height = 22
            ws_an.row_dimensions[row_cur+1].height = 22
            row_cur += 3
            
            # 4. Sub-table Headers Side-by-Side
            # Left table title: Cols A..D
            ws_an.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur, end_column=4)
            cell_lh = ws_an.cell(row=row_cur, column=1, value=f"🏛️ Bút Toán HCM Bị Thiếu ({len(h_missing)} chứng từ)")
            cell_lh.font = font_sub_tbl_title
            cell_lh.alignment = align_left
            
            # Right table title: Cols F..I
            ws_an.merge_cells(start_row=row_cur, start_column=6, end_row=row_cur, end_column=9)
            cell_rh = ws_an.cell(row=row_cur, column=6, value=f"🏢 Bút Toán Điện Lực Bị Thiếu ({len(p_missing)} chứng từ)")
            cell_rh.font = font_sub_tbl_title
            cell_rh.alignment = align_left
            row_cur += 1
            
            # 5. Column Headers for Side-by-Side Tables
            cols_left = ["Ngày CT", "Số CT GL", "Số Tiền (VNĐ)", "Diễn Giải"]
            cols_right = ["Ngày CT", "Số CT GL", "Số Tiền (VNĐ)", "Diễn Giải"]
            
            for c_i, h_text in enumerate(cols_left, 1):
                cell = ws_an.cell(row=row_cur, column=c_i, value=h_text)
                cell.font = font_tbl_header
                cell.fill = fill_tbl_header
                cell.alignment = align_center
                cell.border = border_thin_grid
                
            ws_an.cell(row=row_cur, column=5, value="") # Margin separator column E
            
            for c_i, h_text in enumerate(cols_right, 6):
                cell = ws_an.cell(row=row_cur, column=c_i, value=h_text)
                cell.font = font_tbl_header
                cell.fill = fill_tbl_header
                cell.alignment = align_center
                cell.border = border_thin_grid
            row_cur += 1
            
            # 6. Data Rows Side-by-Side
            max_len = max(len(h_missing), len(p_missing), 1)
            
            for i in range(max_len):
                # Left Table (A..D)
                if i < len(h_missing):
                    u = h_missing[i]
                    c_d = ws_an.cell(row=row_cur, column=1, value=str(u['date'])[:10])
                    c_g = ws_an.cell(row=row_cur, column=2, value=str(u['gl_doc']))
                    c_v = ws_an.cell(row=row_cur, column=3, value=u['net'])
                    c_s = ws_an.cell(row=row_cur, column=4, value=str(u['desc']))
                    
                    c_d.alignment = align_center
                    c_g.alignment = align_center
                    c_v.alignment = align_right
                    c_v.number_format = '#,##0;(#,##0);"-"'
                    c_s.alignment = align_left
                    
                    for c_i in range(1, 5):
                        ws_an.cell(row=row_cur, column=c_i).font = font_tbl_data
                        ws_an.cell(row=row_cur, column=c_i).border = border_thin_grid
                elif len(h_missing) == 0 and i == 0:
                    ws_an.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur, end_column=4)
                    c_ok = ws_an.cell(row=row_cur, column=1, value="✅ Phía HCM đã hạch toán đầy đủ.")
                    c_ok.font = font_success
                    c_ok.fill = fill_success_bg
                    c_ok.alignment = align_left
                    for c_i in range(1, 5):
                        ws_an.cell(row=row_cur, column=c_i).border = border_thin_grid

                # Margin Separator Col E
                ws_an.cell(row=row_cur, column=5, value="")
                
                # Right Table (F..I)
                if i < len(p_missing):
                    u = p_missing[i]
                    c_d = ws_an.cell(row=row_cur, column=6, value=str(u['date'])[:10])
                    c_g = ws_an.cell(row=row_cur, column=7, value=str(u['gl_doc']))
                    c_v = ws_an.cell(row=row_cur, column=8, value=u['net'])
                    c_s = ws_an.cell(row=row_cur, column=9, value=str(u['desc']))
                    
                    c_d.alignment = align_center
                    c_g.alignment = align_center
                    c_v.alignment = align_right
                    c_v.number_format = '#,##0;(#,##0);"-"'
                    c_s.alignment = align_left
                    
                    for c_i in range(6, 10):
                        ws_an.cell(row=row_cur, column=c_i).font = font_tbl_data
                        ws_an.cell(row=row_cur, column=c_i).border = border_thin_grid
                elif len(p_missing) == 0 and i == 0:
                    ws_an.merge_cells(start_row=row_cur, start_column=6, end_row=row_cur, end_column=9)
                    c_ok = ws_an.cell(row=row_cur, column=6, value="✅ Phía Điện lực đã hạch toán đầy đủ.")
                    c_ok.font = font_success
                    c_ok.fill = fill_success_bg
                    c_ok.alignment = align_left
                    for c_i in range(6, 10):
                        ws_an.cell(row=row_cur, column=c_i).border = border_thin_grid
                        
                row_cur += 1

            # Separator rows between pair blocks
            row_cur += 2

        # Set Column Widths for Perfect Side-by-Side Presentation
        ws_an.column_dimensions['A'].width = 14
        ws_an.column_dimensions['B'].width = 13
        ws_an.column_dimensions['C'].width = 22
        ws_an.column_dimensions['D'].width = 45

        ws_an.column_dimensions['E'].width = 4 # Separator column

        ws_an.column_dimensions['F'].width = 14
        ws_an.column_dimensions['G'].width = 13
        ws_an.column_dimensions['H'].width = 22
        ws_an.column_dimensions['I'].width = 45

    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    wb.save(output_filepath)
    try:
        print(f"Exported Excel successfully to: {output_filepath}")
    except Exception:
        pass
    return output_filepath

if __name__ == "__main__":
    from data_processor import process_month_folder, analyze_discrepancy_causes
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    folder_t7 = os.path.join(base_dir, "đầu vào", "tháng 7")
    df_th, h_accs, p_accs, _ = process_month_folder(folder_t7)
    cm, un = analyze_discrepancy_causes(df_th, h_accs, p_accs)
    out_file = os.path.join(base_dir, "đầu ra", "Bao_Cao_Doi_Soat_136_336_Thang_7_2026.xlsx")
    export_reconciliation_excel(df_th, out_file, "Tháng 7/2026", cross_matches=cm, unmatched_by_pair=un)
